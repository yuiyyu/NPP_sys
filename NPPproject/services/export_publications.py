"""
Експорт публікацій у .xlsx файл.
"""
import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


HEADERS = [
    "№",
    "Назва",
    "Тип",
    "Рік",
    "Журнал / видавець",
    "DOI",
    "URL",
    "Автори",
]

COL_WIDTHS = [5, 50, 25, 8, 30, 25, 30, 35]

TYPE_LABELS = {
    "article":          "Стаття",
    "monograph":        "Монографія",
    "conference_paper": "Тези / матеріали конференції",
    "textbook":         "Підручник / навчальний посібник",
    "patent":           "Патент",
    "other":            "Інше",
}

HEADER_FILL  = PatternFill("solid", start_color="2E75B6", end_color="2E75B6")
ALT_FILL     = PatternFill("solid", start_color="DCE6F1", end_color="DCE6F1")
THIN_BORDER  = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def export_publications(rows: list, output_path: str, title: str = "Публікації"):
    """
    rows — список кортежів:
        (id, title, type, year, journal, doi, url, authors)
    output_path — шлях до .xlsx файлу
    title — заголовок аркуша
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Публікації"

    # ── Заголовок документа ───────────────────────────────────────────────────
    ws.merge_cells(f"A1:{get_column_letter(len(HEADERS))}1")
    title_cell = ws["A1"]
    title_cell.value = title
    title_cell.font      = Font(name="Arial", bold=True, size=13, color="FFFFFF")
    title_cell.fill      = PatternFill("solid", start_color="1F3864", end_color="1F3864")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    # Дата генерації
    ws.merge_cells(f"A2:{get_column_letter(len(HEADERS))}2")
    date_cell = ws["A2"]
    date_cell.value = f"Сформовано: {datetime.now().strftime('%d.%m.%Y %H:%M')}"
    date_cell.font      = Font(name="Arial", italic=True, size=10, color="666666")
    date_cell.alignment = Alignment(horizontal="right")
    ws.row_dimensions[2].height = 16

    # ── Заголовки колонок ─────────────────────────────────────────────────────
    for col_i, header in enumerate(HEADERS, start=1):
        cell = ws.cell(row=3, column=col_i, value=header)
        cell.font      = Font(name="Arial", bold=True, size=11, color="FFFFFF")
        cell.fill      = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = THIN_BORDER
    ws.row_dimensions[3].height = 22

    # ── Дані ─────────────────────────────────────────────────────────────────
    for row_i, row in enumerate(rows, start=1):
        excel_row = row_i + 3
        is_alt    = row_i % 2 == 0

        _, title_val, pub_type, year, journal, doi, url, authors = row

        values = [
            row_i,
            title_val or "",
            TYPE_LABELS.get(pub_type, pub_type or ""),
            year or "",
            journal or "",
            doi or "",
            url or "",
            authors or "",
        ]

        for col_i, val in enumerate(values, start=1):
            cell = ws.cell(row=excel_row, column=col_i, value=val)
            cell.font      = Font(name="Arial", size=10)
            cell.border    = THIN_BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if is_alt:
                cell.fill = ALT_FILL
            if col_i == 1:
                cell.alignment = Alignment(horizontal="center", vertical="top")
            if col_i == 4:
                cell.alignment = Alignment(horizontal="center", vertical="top")

        ws.row_dimensions[excel_row].height = 40

    # ── Ширини колонок ────────────────────────────────────────────────────────
    for col_i, width in enumerate(COL_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(col_i)].width = width

    # ── Закріпити рядок заголовків ────────────────────────────────────────────
    ws.freeze_panes = "A4"

    # ── Автофільтр ────────────────────────────────────────────────────────────
    ws.auto_filter.ref = f"A3:{get_column_letter(len(HEADERS))}3"

    wb.save(output_path)
    return output_path