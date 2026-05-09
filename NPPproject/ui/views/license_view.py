from PySide6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout,
    QTableWidget, QTableWidgetItem,
    QPushButton, QHeaderView, QAbstractItemView,
    QLabel, QMessageBox, QFileDialog,
    QListWidget, QListWidgetItem, QSplitter,
    QCheckBox
)
from PySide6.QtCore import Qt, QObject, QEvent
from PySide6.QtGui import QFont, QColor

from services.license_service import (
    get_teachers_for_license_table,
    get_disciplines_by_teacher,
    save_disciplines,
    AUTO_FIELD_KEYS,
)
from ui.dialogs.license_row_dialog import LicenseRowDialog

# ── Колонки таблиці ───────────────────────────────────────────────────────────
# (заголовок, ключ, ширина)
COLUMNS = [
    ("№",                                          "num",          44),
    ("Прізвище, ім'я та по батькові",              "full_name",   200),
    ("Назва освітнього компонента",                "course_name", 220),
    ("Спеціальність",                              "specialty",   110),
    ("Силабус\n(посилання)",                       "syllabus_url",170),
    ("Робоча програма\n(посилання)",               "program_url", 170),
    ("Фахові публікації за ОК",                    "articles",    230),
    ("Тези / матеріали конференцій",               "conferences", 210),
    ("Підручники / посібники",                     "textbooks",   210),
    ("Методичні матеріали",                        "methodical",  200),
    ("Підвищення кваліфікації",                    "trainings",   230),
]

DISC_KEYS = {"course_name", "specialty", "syllabus_url", "program_url"}
AUTO_KEYS = set(AUTO_FIELD_KEYS)

# Кольори
C_NAME   = "#eef2ff"   # світло-синій — ПІБ
C_AUTO   = "#f0f0f0"   # сірий        — дані з БД
C_EDITED = "#e8f4e8"   # зелений      — відредаговано вручну
C_FILLED = "#ffffff"   # білий        — ручне поле заповнено
C_EMPTY  = "#fffbe6"   # жовтий       — ручне поле порожнє
C_TEXT   = "#000000"


def _item(text: str, bg: str, align=None) -> QTableWidgetItem:
    it = QTableWidgetItem(str(text))
    it.setBackground(QColor(bg))
    it.setForeground(QColor(C_TEXT))
    it.setFlags(it.flags() & ~Qt.ItemIsEditable)
    it.setTextAlignment(align or (Qt.AlignLeft | Qt.AlignTop))
    return it


class _WheelFilter(QObject):
    def __init__(self, table):
        super().__init__(table)
        self._t = table

    def eventFilter(self, obj, event):
        if obj is self._t.viewport() and event.type() == QEvent.Wheel:
            if event.modifiers() & Qt.ShiftModifier:
                bar = self._t.horizontalScrollBar()
                bar.setValue(bar.value() - event.angleDelta().y())
                return True
        return super().eventFilter(obj, event)


class LicenseView(QWidget):
    def __init__(self):
        super().__init__()
        self._teachers = []          # [{teacher_id, full_name, articles, ...}]
        self._cur_tid  = None        # поточний teacher_id

        lay = QVBoxLayout(self)
        lay.setContentsMargins(12, 12, 12, 12)
        lay.setSpacing(8)

        # ── Панель кнопок ─────────────────────────────────────────────────────
        btn_row = QHBoxLayout()
        self._btn_refresh = QPushButton("🔄 Оновити з БД")
        self._btn_edit    = QPushButton("✏️ Редагувати викладача")
        self._btn_clear   = QPushButton("🗑 Очистити дисципліни")
        self._btn_exp_one = QPushButton("📄 Експорт (поточний)")
        self._btn_exp_all = QPushButton("📊 Експорт (всі)")
        for b in [self._btn_refresh, self._btn_edit, self._btn_clear,
                  self._btn_exp_one, self._btn_exp_all]:
            btn_row.addWidget(b)
        btn_row.addStretch()
        self._chk_5y = QCheckBox("📅 Тільки останні 5 років")
        self._chk_5y.setStyleSheet("font-size:11px; color:#2c3e50;")
        btn_row.addWidget(self._chk_5y)
        lay.addLayout(btn_row)

        # ── Сплітер ───────────────────────────────────────────────────────────
        splitter = QSplitter(Qt.Horizontal)

        # Ліва: список викладачів
        lp = QWidget()
        ll = QVBoxLayout(lp)
        ll.setContentsMargins(0, 0, 4, 0)
        ll.setSpacing(4)
        lbl_t = QLabel("Викладачі:")
        bf2 = QFont(); bf2.setBold(True)
        lbl_t.setFont(bf2)
        ll.addWidget(lbl_t)
        self._teacher_list = QListWidget()
        self._teacher_list.setFixedWidth(230)
        ll.addWidget(self._teacher_list)

        # Права: поточний викладач + таблиця
        rp = QWidget()
        rl = QVBoxLayout(rp)
        rl.setContentsMargins(4, 0, 0, 0)
        rl.setSpacing(4)
        self._cur_label = QLabel("← Оберіть викладача зі списку")
        self._cur_label.setStyleSheet("color:#888; font-style:italic;")
        rl.addWidget(self._cur_label)

        self._table = QTableWidget()
        self._table.setColumnCount(len(COLUMNS))
        self._table.setHorizontalHeaderLabels([c[0] for c in COLUMNS])
        self._table.setEditTriggers(QTableWidget.NoEditTriggers)
        self._table.setSelectionBehavior(QTableWidget.SelectRows)
        self._table.setAlternatingRowColors(False)
        self._table.verticalHeader().setVisible(False)
        self._table.verticalHeader().setDefaultSectionSize(80)
        self._table.setWordWrap(True)
        self._table.setHorizontalScrollMode(QAbstractItemView.ScrollPerPixel)
        self._table.setVerticalScrollMode(QAbstractItemView.ScrollPerPixel)
        self._table.setStyleSheet("""
            QTableWidget::item:selected { background:#cfe2ff; color:#000000; }
            QTableWidget::item:selected:!active { background:#ddeeff; color:#000000; }
        """)
        hdr = self._table.horizontalHeader()
        hdr.setDefaultAlignment(Qt.AlignCenter)
        for i, (_, _, w) in enumerate(COLUMNS):
            hdr.setSectionResizeMode(i, QHeaderView.Interactive)
            self._table.setColumnWidth(i, w)
        self._wheel = _WheelFilter(self._table)
        self._table.viewport().installEventFilter(self._wheel)
        rl.addWidget(self._table)

        splitter.addWidget(lp)
        splitter.addWidget(rp)
        splitter.setSizes([230, 900])
        splitter.setStretchFactor(0, 0)
        splitter.setStretchFactor(1, 1)
        lay.addWidget(splitter, stretch=1)

        # ── Сигнали ───────────────────────────────────────────────────────────
        self._btn_refresh.clicked.connect(self.load_data)
        self._btn_edit.clicked.connect(self._open_dialog)
        self._btn_clear.clicked.connect(self._clear_disciplines)
        self._btn_exp_one.clicked.connect(self._export_current)
        self._btn_exp_all.clicked.connect(self._export_all)
        self._chk_5y.stateChanged.connect(self.load_data)
        self._teacher_list.currentRowChanged.connect(self._on_teacher_changed)
        self._table.doubleClicked.connect(self._open_dialog)

        self.load_data()

    # ── Завантаження ──────────────────────────────────────────────────────────

    def load_data(self):
        self._teachers = get_teachers_for_license_table(
            last_5_years=self._chk_5y.isChecked()
        )
        self._fill_teacher_list()
        # Відновлюємо вибір
        if self._cur_tid:
            for i, t in enumerate(self._teachers):
                if t["teacher_id"] == self._cur_tid:
                    self._teacher_list.setCurrentRow(i)
                    return
        if self._teachers:
            self._teacher_list.setCurrentRow(0)

    def _fill_teacher_list(self):
        self._teacher_list.blockSignals(True)
        self._teacher_list.clear()
        for t in self._teachers:
            item = QListWidgetItem(t["full_name"])
            item.setData(Qt.UserRole, t["teacher_id"])
            self._teacher_list.addItem(item)
        self._teacher_list.blockSignals(False)

    # ── Вибір викладача ───────────────────────────────────────────────────────

    def _on_teacher_changed(self, idx: int):
        if idx < 0 or idx >= len(self._teachers):
            return
        t = self._teachers[idx]
        self._cur_tid = t["teacher_id"]
        self._cur_label.setText(f"{t['full_name']}")
        self._cur_label.setStyleSheet("color:#222; font-weight:bold;")
        self._fill_table(idx)

    def _fill_table(self, teacher_idx: int):
        t     = self._teachers[teacher_idx]
        tid   = t["teacher_id"]
        discs = get_disciplines_by_teacher(tid, last_5_years=self._chk_5y.isChecked())

        rows = max(len(discs), 1)
        self._table.setRowCount(rows)

        for di in range(rows):
            disc     = discs[di] if di < len(discs) else {}
            is_first = (di == 0)

            for ci, (_, key, _) in enumerate(COLUMNS):

                if key == "num":
                    cell = _item(str(di + 1), C_AUTO, Qt.AlignCenter)

                elif key == "full_name":
                    cell = _item(t["full_name"] if is_first else "", C_NAME)

                elif key in DISC_KEYS:
                    val  = (disc.get(key) or "").strip()
                    cell = _item(val, C_FILLED if val else C_EMPTY)

                elif key in AUTO_KEYS:
                    raw_type = disc.get(f"_raw_{key}", "")  # тип: "", "__DB__", JSON, текст
                    display  = (disc.get(key) or "").strip()  # вже розкритий текст

                    if not raw_type:
                        # Порожньо
                        cell = _item("", C_EMPTY if not is_first else C_AUTO)
                    elif raw_type == "__DB__":
                        # Всі з БД — сірий
                        cell = _item(display, C_AUTO)
                    elif raw_type.startswith("{"):
                        # JSON ids — вибрано вручну, зелений
                        cell = _item(display, C_EDITED)
                    else:
                        # Вручну введений текст — зелений
                        cell = _item(display, C_EDITED)

                else:
                    cell = _item("", C_AUTO)

                self._table.setItem(di, ci, cell)

        self._table.resizeRowsToContents()

    # ── Діалог редагування ────────────────────────────────────────────────────

    def _open_dialog(self):
        idx = self._teacher_list.currentRow()
        if idx < 0 or idx >= len(self._teachers):
            QMessageBox.information(self, "Увага", "Оберіть викладача зі списку.")
            return
        t     = self._teachers[idx]
        tid   = t["teacher_id"]
        discs = get_disciplines_by_teacher(tid, last_5_years=self._chk_5y.isChecked())

        dlg = LicenseRowDialog(
            parent=self,
            row_data=t,
            disciplines=discs,
            last_5_years=self._chk_5y.isChecked(),
        )
        dlg.saved.connect(lambda new_discs, t_id=tid, t_idx=idx:
                          self._on_saved(t_id, t_idx, new_discs))
        dlg.exec()

    def _on_saved(self, teacher_id: str, teacher_idx: int, disciplines: list):
        try:
            save_disciplines(teacher_id, disciplines)
        except Exception as e:
            QMessageBox.critical(self, "Помилка збереження", str(e))
            return
        # Перезавантажуємо весь список з БД
        self._teachers = get_teachers_for_license_table(
            last_5_years=self._chk_5y.isChecked()
        )
        # Шукаємо актуальний індекс по teacher_id (список міг перебудуватись)
        new_idx = next(
            (i for i, t in enumerate(self._teachers) if t["teacher_id"] == teacher_id),
            teacher_idx
        )
        self._cur_tid = teacher_id
        self._fill_teacher_list()
        self._teacher_list.blockSignals(True)
        self._teacher_list.setCurrentRow(new_idx)
        self._teacher_list.blockSignals(False)
        self._fill_table(new_idx)

    # ── Очищення ──────────────────────────────────────────────────────────────

    def _clear_disciplines(self):
        idx = self._teacher_list.currentRow()
        if idx < 0:
            QMessageBox.information(self, "Увага", "Оберіть викладача.")
            return
        t = self._teachers[idx]
        if QMessageBox.question(
            self, "Підтвердження",
            f"Видалити всі дисципліни для {t['full_name']}?",
            QMessageBox.Yes | QMessageBox.No
        ) == QMessageBox.Yes:
            try:
                save_disciplines(t["teacher_id"], [])
            except Exception as e:
                QMessageBox.critical(self, "Помилка", str(e))
                return
            self._fill_table(idx)

    # ── Експорт ───────────────────────────────────────────────────────────────

    def _export_current(self):
        idx = self._teacher_list.currentRow()
        if idx < 0:
            QMessageBox.information(self, "Увага", "Оберіть викладача.")
            return
        t = self._teachers[idx]
        self._do_export([t], f"license_{t['full_name'].split()[0]}.xlsx")

    def _export_all(self):
        self._do_export(self._teachers, "license_table_all.xlsx")

    def _do_export(self, teachers: list, default_name: str):
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        except ImportError:
            QMessageBox.critical(self, "Помилка",
                "Виконайте: pip install openpyxl")
            return

        path, _ = QFileDialog.getSaveFileName(
            self, "Зберегти Excel", default_name, "Excel Files (*.xlsx)")
        if not path:
            return

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Самооцінювання"

        thin   = Side(style="thin")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        def fill(hex6): return PatternFill("solid", fgColor=hex6)
        def al(h="left", v="top", wrap=True):
            return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

        hfill = fill("D9E1F2")
        nc    = len(COLUMNS)
        lc    = openpyxl.utils.get_column_letter(nc)

        # Рядок 1: кафедра
        ws.merge_cells(f"A1:{lc}1")
        ws["A1"] = "Кафедра: _______________________________________________"
        ws["A1"].font = Font(bold=True, size=11)
        ws["A1"].alignment = al("left", "center")
        ws.row_dimensions[1].height = 22

        # Рядок 2: заголовок
        ws.merge_cells(f"A2:{lc}2")
        ws["A2"] = (
            "САМООЦІНЮВАННЯ\n"
            "відповідності публікацій науково-педагогічних (педагогічних) "
            "працівників змісту освітніх компонентів, які вони викладають"
        )
        ws["A2"].font = Font(bold=True, size=11)
        ws["A2"].alignment = al("center", "center")
        ws.row_dimensions[2].height = 50

        # Рядок 3: заголовки колонок
        for ci, (label, _, _) in enumerate(COLUMNS, 1):
            c = ws.cell(row=3, column=ci, value=label)
            c.font = Font(bold=True, size=10)
            c.alignment = al("center", "center")
            c.fill = hfill
            c.border = border
        ws.row_dimensions[3].height = 54

        # Ширини колонок
        col_w = [6, 26, 28, 13, 22, 22, 32, 28, 28, 26, 32]
        for i, w in enumerate(col_w[:nc], 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

        excel_row = 4
        global_n  = 1

        for t in teachers:
            tid   = t["teacher_id"]
            discs = get_disciplines_by_teacher(tid, last_5_years=self._chk_5y.isChecked())
            rows  = max(len(discs), 1)

            for di in range(rows):
                disc     = discs[di] if di < len(discs) else {}
                is_first = (di == 0)

                for ci, (_, key, _) in enumerate(COLUMNS, 1):
                    if key == "num":
                        val, bg, a = str(global_n), "F0F0F0", al("center", "center")

                    elif key == "full_name":
                        val = t["full_name"] if is_first else ""
                        bg, a = "EEF2FF", al("left", "top")

                    elif key in DISC_KEYS:
                        val = (disc.get(key) or "").strip()
                        bg  = "FFFFFF" if val else "FFFBE6"
                        a   = al("left", "top")

                    elif key in AUTO_KEYS:
                        raw_type = disc.get(f"_raw_{key}", "")
                        display  = (disc.get(key) or "").strip()
                        if not raw_type:
                            val, bg = "", "FFFBE6"
                        elif raw_type == "__DB__":
                            val, bg = display, "F0F0F0"
                        else:
                            val, bg = display, "E8F4E8"
                        a = al("left", "top")

                    else:
                        val, bg, a = "", "F0F0F0", al("left", "top")

                    c = ws.cell(row=excel_row, column=ci, value=val)
                    c.font = Font(size=9, color="000000")
                    c.alignment = a
                    c.fill = fill(bg)
                    c.border = border

                ws.row_dimensions[excel_row].height = 62
                excel_row += 1
                global_n  += 1

        try:
            wb.save(path)
            QMessageBox.information(self, "Готово", f"Збережено:\n{path}")
            import os
            if os.name == "nt":
                os.startfile(path)
        except Exception as e:
            QMessageBox.critical(self, "Помилка збереження", str(e))